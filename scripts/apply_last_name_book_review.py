#!/usr/bin/env python3
"""Validate and apply the completed last-name discrepancy review workbook.

The stage preserves the source row order and schema. It applies only explicit
owner inputs from columns A:D of the Discrepancies worksheet and never edits a
canonical file in place.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


DECISIONS = {"Keep staged", "Correct manually"}
INPUT_COLUMNS = [
    "review_decision",
    "corrected_last_name",
    "alias_name_action",
    "review_notes",
]
PROTECTED_COLUMNS = [
    "review_priority",
    "review_category",
    "person_id",
    "name_raw",
    "staged_first_name",
    "staged_patronymic_name",
    "staged_last_name",
    "book_last_name",
    "book_name_field",
    "similarity",
    "discrepancy_type",
    "book_name_parsing",
    "settlement",
    "book_settlement",
    "person_order_in_settlement",
    "source_position_id",
]
FIRST_NAME_NOTE = re.compile(r"add first_name\s+`([^`]+)`", re.IGNORECASE)
REMOVE_PATRONYMIC_NOTE = re.compile(
    r"remove\s+staged_patronymic_name", re.IGNORECASE
)


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def read_review(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Discrepancies" not in workbook.sheetnames:
        raise ValueError("review workbook is missing the Discrepancies worksheet")
    sheet = workbook["Discrepancies"]
    raw_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    header_index = next(
        (i for i, row in enumerate(raw_rows) if clean(row[0]) == "review_decision"),
        None,
    )
    if header_index is None:
        raise ValueError("review workbook header was not found")
    headers = [clean(value) for value in raw_rows[header_index]]
    required = set(INPUT_COLUMNS + PROTECTED_COLUMNS)
    missing = required - set(headers)
    if missing:
        raise ValueError("review workbook is missing columns: " + ", ".join(sorted(missing)))
    records = []
    for values in raw_rows[header_index + 1 :]:
        row = {header: clean(value) for header, value in zip(headers, values)}
        if row["person_id"]:
            records.append(row)
    return headers, records


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--decisions-output", type=Path, required=True)
    parser.add_argument("--diff-output", type=Path, required=True)
    parser.add_argument("--qa-output", type=Path, required=True)
    args = parser.parse_args()

    fields, source_rows = read_csv(args.source)
    required_source = {
        "person_id", "source_position_id", "name_raw", "name_alias",
        "first_name", "patronymic_name", "last_name",
        "first_name_source", "patronymic_source", "last_name_source",
        "parse_status", "parse_confidence", "parse_rule",
        "name_order_detected", "manual_review_reason",
    }
    missing_source = required_source - set(fields)
    if missing_source:
        raise ValueError("source is missing fields: " + ", ".join(sorted(missing_source)))

    _, reviews = read_review(args.workbook)
    if len(reviews) != 249:
        raise ValueError(f"expected 249 review records, found {len(reviews)}")
    review_by_id = {row["person_id"]: row for row in reviews}
    if len(review_by_id) != len(reviews):
        raise ValueError("review workbook contains duplicate person_id values")
    invalid = [row["person_id"] for row in reviews if row["review_decision"] not in DECISIONS]
    if invalid:
        raise ValueError("blank or unsupported decisions for: " + ", ".join(invalid))

    source_by_id = {row["person_id"]: row for row in source_rows}
    missing_ids = sorted(set(review_by_id) - set(source_by_id))
    if missing_ids:
        raise ValueError("review IDs absent from source: " + ", ".join(missing_ids))

    protected_mismatches = []
    for person_id, review in review_by_id.items():
        source = source_by_id[person_id]
        expected = {
            "person_id": source["person_id"],
            "name_raw": source["name_raw"],
            "staged_first_name": source["first_name"],
            "staged_patronymic_name": source["patronymic_name"],
            "staged_last_name": source["last_name"],
            "settlement": source["settlement"],
            "person_order_in_settlement": source["person_order_in_settlement"],
            "source_position_id": source["source_position_id"],
        }
        for field, value in expected.items():
            if review[field] != value:
                protected_mismatches.append(
                    {"person_id": person_id, "field": field, "expected": value, "actual": review[field]}
                )
    if protected_mismatches:
        raise ValueError(f"protected workbook/source mismatches: {protected_mismatches[:10]}")

    decisions = []
    output_rows = []
    diffs = []
    changed_ids = set()
    alias_additions = 0
    component_changes = Counter()
    confirmed_noop_ids = []

    for before in source_rows:
        person_id = before["person_id"]
        review = review_by_id.get(person_id)
        after = dict(before)
        if review is not None:
            decisions.append({field: review[field] for field in INPUT_COLUMNS + PROTECTED_COLUMNS})
        if review is None or review["review_decision"] == "Keep staged":
            output_rows.append(after)
            continue

        corrected_last = review["corrected_last_name"]
        alias = review["alias_name_action"]
        notes = review["review_notes"]
        first_match = FIRST_NAME_NOTE.search(notes)
        remove_patronymic = bool(REMOVE_PATRONYMIC_NOTE.search(notes))
        if not any((corrected_last, alias, first_match, remove_patronymic)):
            raise ValueError(f"Correct manually has no actionable input for {person_id}")

        if alias:
            if after["name_alias"] and after["name_alias"] != alias:
                raise ValueError(
                    f"{person_id} would overwrite existing name_alias {after['name_alias']!r} with {alias!r}"
                )
            if after["name_alias"] != alias:
                after["name_alias"] = alias
                alias_additions += 1

        components_changed = False
        if first_match:
            new_first = first_match.group(1).strip()
            if after["first_name"] != new_first:
                after["first_name"] = new_first
                after["first_name_source"] = "reviewed_exception"
                component_changes["first_name"] += 1
                components_changed = True
        if remove_patronymic and after["patronymic_name"]:
            after["patronymic_name"] = ""
            after["patronymic_source"] = ""
            component_changes["patronymic_name"] += 1
            components_changed = True
        if corrected_last and after["last_name"] != corrected_last:
            after["last_name"] = corrected_last
            after["last_name_source"] = "reviewed_exception"
            component_changes["last_name"] += 1
            components_changed = True
        if components_changed:
            after["parse_status"] = "observed"
            after["parse_confidence"] = "high"
            after["parse_rule"] = "reviewed_last_name_book_index"
            after["name_order_detected"] = "reviewed_exception"
            after["manual_review_reason"] = (
                "Owner-approved last-name book discrepancy review"
                + (f": {notes}" if notes else "")
            )

        changed_fields = [field for field in fields if before[field] != after[field]]
        if not changed_fields:
            confirmed_noop_ids.append(person_id)
            output_rows.append(after)
            continue
        changed_ids.add(person_id)
        for field in changed_fields:
            diffs.append({
                "person_id": person_id,
                "source_position_id": before["source_position_id"],
                "field": field,
                "before": before[field],
                "after": after[field],
                "review_decision": review["review_decision"],
                "review_notes": notes,
            })
        output_rows.append(after)

    if len(changed_ids) + len(confirmed_noop_ids) != 49:
        raise ValueError(
            "expected 49 manual decisions, found "
            f"{len(changed_ids)} changed and {len(confirmed_noop_ids)} confirmed no-op records"
        )

    allowed_change_fields = {
        "name_alias", "first_name", "patronymic_name", "last_name",
        "first_name_source", "patronymic_source", "last_name_source",
        "parse_status", "parse_confidence", "parse_rule",
        "name_order_detected", "manual_review_reason",
    }
    unexpected_fields = sorted({row["field"] for row in diffs} - allowed_change_fields)
    if unexpected_fields:
        raise ValueError("unexpected changed fields: " + ", ".join(unexpected_fields))

    p5626 = next(row for row in output_rows if row["person_id"] == "P005626")
    if (p5626["first_name"], p5626["patronymic_name"], p5626["last_name"]) != (
        "Рарица", "", "Марина"
    ):
        raise ValueError("P005626 did not resolve to first_name=Рарица, last_name=Марина")

    write_csv(args.output, output_rows, fields)
    write_csv(args.decisions_output, decisions, INPUT_COLUMNS + PROTECTED_COLUMNS)
    write_csv(
        args.diff_output,
        diffs,
        ["person_id", "source_position_id", "field", "before", "after", "review_decision", "review_notes"],
    )

    qa = {
        "status": "passed",
        "source": str(args.source).replace("\\", "/"),
        "source_sha256": sha256(args.source),
        "review_workbook": str(args.workbook).replace("\\", "/"),
        "review_workbook_sha256": sha256(args.workbook),
        "staged_output": str(args.output).replace("\\", "/"),
        "staged_output_sha256": sha256(args.output),
        "records": len(output_rows),
        "columns": len(fields),
        "unique_person_ids": len({row["person_id"] for row in output_rows}),
        "review_records": len(reviews),
        "decision_counts": dict(sorted(Counter(row["review_decision"] for row in reviews).items())),
        "changed_records": len(changed_ids),
        "confirmed_noop_records": len(confirmed_noop_ids),
        "confirmed_noop_person_ids": confirmed_noop_ids,
        "changed_cells": len(diffs),
        "alias_additions": alias_additions,
        "component_changes": dict(sorted(component_changes.items())),
        "protected_workbook_mismatches": 0,
        "unexpected_changed_fields": unexpected_fields,
        "row_order_preserved": [row["person_id"] for row in source_rows]
        == [row["person_id"] for row in output_rows],
        "schema_preserved": True,
        "p005626": {
            "first_name": p5626["first_name"],
            "patronymic_name": p5626["patronymic_name"],
            "last_name": p5626["last_name"],
        },
        "canonical_data_modified": False,
    }
    args.qa_output.parent.mkdir(parents=True, exist_ok=True)
    args.qa_output.write_text(
        json.dumps(qa, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(qa, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
