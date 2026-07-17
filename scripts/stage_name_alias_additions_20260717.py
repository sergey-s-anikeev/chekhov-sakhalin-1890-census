import csv
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data/review/alias_name_add_20260717/Sakhalin_1890_name_alias_update.xlsx"
SOURCE = ROOT / "data/staging/arrival_year_item21_20260717/clean_sakhalin_1890_ru_v3_20260712_items7_8_item12_item16_item18_comments_item21_staged.csv"
REVIEW = ROOT / "data/review/alias_name_add_20260717/owner_response"
STAGE = ROOT / "data/staging/name_alias_add_20260717"
QA = ROOT / "outputs/qa/name_alias_add_20260717"
OUTPUT = STAGE / "clean_sakhalin_1890_ru_v3_20260712_items7_8_item12_item16_item18_comments_item21_alias_staged.csv"


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def read_alias_workbook():
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["clean_sakhalin_1890_ru_v3_20260"]
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator)
    index = {name: i for i, name in enumerate(header)}
    required = {"person_id", "name_raw", "name_alias"}
    if not required.issubset(index):
        raise ValueError(f"Workbook sheet is missing required columns: {sorted(required - set(index))}")
    records = {}
    for values in iterator:
        person_id = "" if values[index["person_id"]] is None else str(values[index["person_id"]])
        if not person_id:
            continue
        if person_id in records:
            raise ValueError(f"Duplicate person_id in workbook: {person_id}")
        records[person_id] = {
            "name_raw": "" if values[index["name_raw"]] is None else str(values[index["name_raw"]]),
            "name_alias": "" if values[index["name_alias"]] is None else str(values[index["name_alias"]]),
        }
    workbook.close()
    return records


def main():
    workbook_rows = read_alias_workbook()
    with SOURCE.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames
        source_rows = list(reader)
    if len(workbook_rows) != len(source_rows):
        raise ValueError(f"Workbook/source row-count mismatch: {len(workbook_rows)} vs {len(source_rows)}")

    staged, diff = [], []
    name_mismatches = []
    existing_alias_overwrites = []
    for before in source_rows:
        person_id = before["person_id"]
        if person_id not in workbook_rows:
            raise ValueError(f"Missing workbook person_id: {person_id}")
        feedback = workbook_rows[person_id]
        if before["name_raw"] != feedback["name_raw"]:
            name_mismatches.append(person_id)
        alias_after = feedback["name_alias"]
        if before["name_alias"] and before["name_alias"] != alias_after:
            existing_alias_overwrites.append(person_id)
        after = dict(before)
        after["name_alias"] = alias_after
        staged.append(after)
        if before["name_alias"] != alias_after:
            diff.append({
                "person_id": person_id,
                "page_number": before.get("page_number", ""),
                "name_raw": before["name_raw"],
                "name_alias_before": before["name_alias"],
                "name_alias_after": alias_after,
            })
    if name_mismatches:
        raise ValueError(f"name_raw mismatch for {len(name_mismatches)} records")
    if existing_alias_overwrites:
        raise ValueError(f"Workbook would overwrite {len(existing_alias_overwrites)} existing aliases")
    if len(diff) != 14:
        raise ValueError(f"Expected 14 alias additions, found {len(diff)}")

    write_csv(OUTPUT, staged, fields)
    write_csv(REVIEW / "name_alias_additions_approved.csv", diff, list(diff[0]))
    write_csv(QA / "name_alias_additions_diff.csv", diff, list(diff[0]))

    non_target_changes = 0
    for before, after in zip(source_rows, staged):
        for field in fields:
            if field != "name_alias" and before[field] != after[field]:
                non_target_changes += 1
    qa = {
        "source": str(SOURCE.relative_to(ROOT)),
        "feedback_workbook": str(WORKBOOK.relative_to(ROOT)),
        "staged_output": str(OUTPUT.relative_to(ROOT)),
        "row_count": len(staged),
        "column_count": len(fields),
        "unique_person_ids": len({r["person_id"] for r in staged}),
        "workbook_person_ids": len(workbook_rows),
        "name_raw_mismatches": 0,
        "alias_additions": len(diff),
        "existing_alias_overwrites": 0,
        "populated_name_alias_before": sum(bool(r["name_alias"]) for r in source_rows),
        "populated_name_alias_after": sum(bool(r["name_alias"]) for r in staged),
        "non_target_changes": non_target_changes,
        "canonical_data_modified": False,
    }
    QA.mkdir(parents=True, exist_ok=True)
    (QA / "name_alias_additions_qa.json").write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(qa, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
